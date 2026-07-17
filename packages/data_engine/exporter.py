"""Safe professional Excel evidence-pack exporter."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

import polars as pl
import xlsxwriter
from openpyxl import load_workbook

from packages.contracts import OperationMetric, ValidationFinding, WorkflowConfiguration
from packages.data_engine.safety import SourceFile

SHEETS = [
    "Processed Data",
    "Summary",
    "Validation Errors",
    "Rejected Rows",
    "Run Audit",
    "Source Metadata",
    "Applied Rules",
]


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _write_table(
    workbook: xlsxwriter.Workbook,
    name: str,
    rows: list[dict[str, Any]],
    header_format: Any,
    error_format: Any | None = None,
) -> None:
    worksheet = workbook.add_worksheet(name)
    if not rows:
        worksheet.write(0, 0, "No records", header_format)
        worksheet.set_column(0, 0, 24)
        return
    headers = list(rows[0])
    for column, header in enumerate(headers):
        worksheet.write(0, column, header, header_format)
        values = [str(row.get(header, "")) for row in rows[:1000]]
        width = min(45, max(12, len(header) + 2, *(len(value) + 2 for value in values)))
        worksheet.set_column(column, column, width)
    for row_index, row in enumerate(rows, start=1):
        for column, header in enumerate(headers):
            cell_format = error_format if error_format and name == "Validation Errors" else None
            worksheet.write(row_index, column, _safe_cell(row.get(header)), cell_format)
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, len(rows), len(headers) - 1)


def export_workbook(
    output_directory: Path,
    run_id: UUID,
    source: SourceFile,
    workflow: WorkflowConfiguration,
    processed: pl.DataFrame,
    rejected: list[dict[str, Any]],
    findings: list[ValidationFinding],
    metrics: list[OperationMetric],
    counts: dict[str, int],
) -> Path:
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    filename = f"{workflow.export.filename_prefix}_{timestamp}_{str(run_id)[:8]}.xlsx"
    destination = (output_directory / filename).resolve()
    if destination.exists():
        raise FileExistsError("OUTPUT_COLLISION")
    workbook = xlsxwriter.Workbook(destination, {"constant_memory": True, "strings_to_formulas": False})
    header = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#17324D", "border": 1})
    error = workbook.add_format({"bg_color": "#FDE8E7", "font_color": "#9B1C1C"})
    try:
        _write_table(workbook, "Processed Data", processed.to_dicts(), header)
        summary: list[dict[str, Any]] = [
            {"Metric": "Run ID", "Value": str(run_id)},
            {"Metric": "Workflow", "Value": workflow.display_name},
            {"Metric": "Workflow version", "Value": workflow.workflow_version},
            {"Metric": "Source fingerprint", "Value": source.sha256},
            *({"Metric": key.replace("_", " ").title(), "Value": value} for key, value in counts.items()),
        ]
        _write_table(workbook, "Summary", summary, header)
        _write_table(workbook, "Validation Errors", [item.model_dump(mode="json") for item in findings], header, error)
        _write_table(workbook, "Rejected Rows", rejected, header)
        audit: list[dict[str, Any]] = [
            {"Field": "run_id", "Value": str(run_id)},
            {"Field": "workflow_id", "Value": str(workflow.id)},
            {"Field": "workflow_version", "Value": workflow.workflow_version},
            {"Field": "source_sha256", "Value": source.sha256},
            {"Field": "generated_at_utc", "Value": datetime.now(UTC).isoformat()},
            {"Field": "reconciliation", "Value": json.dumps(counts, sort_keys=True)},
        ]
        _write_table(workbook, "Run Audit", audit, header)
        source_meta: list[dict[str, Any]] = [
            {"Field": "original_filename", "Value": source.original_filename},
            {"Field": "size_bytes", "Value": source.size_bytes},
            {"Field": "sha256", "Value": source.sha256},
        ]
        _write_table(workbook, "Source Metadata", source_meta, header)
        rules: list[dict[str, Any]] = [
            {
                "Kind": "operation",
                "ID": metric.operation_id,
                "Version": metric.operation_version,
                "Affected rows": metric.affected_rows,
                "Rows in": metric.rows_in,
                "Rows out": metric.rows_out,
            }
            for metric in metrics
        ] + [
            {
                "Kind": "validation",
                "ID": rule.id,
                "Version": 1,
                "Affected rows": sum(item.rule_identifier == rule.id for item in findings),
                "Rows in": counts["rows_read"],
                "Rows out": counts["rows_written"],
            }
            for rule in workflow.validation_rules
        ]
        _write_table(workbook, "Applied Rules", rules, header)
    finally:
        workbook.close()
    check = load_workbook(destination, read_only=True, data_only=False)
    try:
        missing = set(SHEETS) - set(check.sheetnames)
        if missing:
            raise RuntimeError(f"OUTPUT_VERIFICATION_FAILED: missing sheets {sorted(missing)}")
    finally:
        check.close()
    return destination
