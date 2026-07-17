"""Generic CSV/Excel inspection without fixed ranges or business fields."""

from __future__ import annotations

import csv
import re
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from packages.contracts import (
    CanonicalType,
    ColumnProfile,
    DiscoveryOverrides,
    DiscoveryResult,
    HeaderCandidate,
    RowClassification,
    SourceHandle,
    TableDiscovery,
)
from packages.data_engine.safety import SourceFile

NULL_LIKE = {"", "null", "none", "n/a", "na", "-"}
INTEGER = re.compile(r"^[+-]?\d+$")
DECIMAL = re.compile(r"^[+-]?(?:\d+\.\d+|\d{1,3}(?:,\d{3})+(?:\.\d+)?)$")
DATE_ISO = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$")
DATE_AMBIGUOUS = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$")
PERCENT = re.compile(r"^[+-]?\d+(?:\.\d+)?%$")

RowKind = Literal[
    "data", "repeated_header", "grand_total", "subtotal", "generated_footer", "signature", "note"
]


@dataclass(slots=True)
class RawSheet:
    name: str
    state: str
    rows: list[list[Any]]
    merged_ranges: list[tuple[int, int, int, int]] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class Region:
    start_row: int
    end_row: int
    start_column: int
    end_column: int

    @property
    def region_name(self) -> str:
        return (
            f"{get_column_letter(self.start_column)}{self.start_row}:"
            f"{get_column_letter(self.end_column)}{self.end_row}"
        )


def _string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalise_width(rows: list[list[Any]]) -> list[list[Any]]:
    width = max((len(row) for row in rows), default=0)
    return [row + [None] * (width - len(row)) for row in rows]


def _load_csv(path: Path, row_limit: int | None = None) -> list[RawSheet]:
    if path.stat().st_size == 0:
        return [RawSheet(path.stem, "visible", [])]
    encoding = "utf-8-sig"
    try:
        with path.open("r", encoding=encoding, newline="") as stream:
            sample = stream.read(8192)
    except UnicodeDecodeError:
        encoding = "cp1252"
        with path.open("r", encoding=encoding, newline="") as stream:
            sample = stream.read(8192)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows: list[list[str]] = []
    with path.open("r", encoding=encoding, newline="") as stream:
        for index, row in enumerate(csv.reader(stream, dialect)):
            if row_limit is not None and index >= row_limit:
                break
            rows.append(list(row))
    return [RawSheet(path.stem, "visible", _normalise_width(rows))]


def _load_excel(path: Path) -> list[RawSheet]:
    # Normal mode is intentional for merged-range metadata. The performance policy
    # warns before large workbooks; this is not represented as streaming ingestion.
    workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    sheets: list[RawSheet] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            merged = [
                (item.min_row, item.max_row, item.min_col, item.max_col)
                for item in worksheet.merged_cells.ranges
            ]
            sheets.append(
                RawSheet(worksheet.title, worksheet.sheet_state, _normalise_width(rows), merged)
            )
    finally:
        workbook.close()
    return sheets


def load_sheets(source: SourceFile) -> list[RawSheet]:
    suffix = source.path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(source.path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel(source.path)
    raise ValueError(f"SOURCE_UNSUPPORTED_TYPE: {suffix}")


def _is_blank(row: Iterable[Any]) -> bool:
    return not any(_string(value) for value in row)


def _blocks(indices: list[int]) -> list[tuple[int, int]]:
    if not indices:
        return []
    result: list[tuple[int, int]] = []
    start = previous = indices[0]
    for current in indices[1:]:
        if current != previous + 1:
            result.append((start, previous))
            start = current
        previous = current
    result.append((start, previous))
    return result


def _regions(rows: list[list[Any]]) -> list[Region]:
    """Find rectangular candidates separated by fully blank rows or columns."""
    active_rows = [index for index, row in enumerate(rows, start=1) if not _is_blank(row)]
    candidates: list[Region] = []
    for row_start, row_end in _blocks(active_rows):
        width = max((len(rows[index - 1]) for index in range(row_start, row_end + 1)), default=0)
        active_columns = [
            column
            for column in range(1, width + 1)
            if any(
                column <= len(rows[index - 1]) and bool(_string(rows[index - 1][column - 1]))
                for index in range(row_start, row_end + 1)
            )
        ]
        for column_start, column_end in _blocks(active_columns):
            populated_rows = [
                index
                for index in range(row_start, row_end + 1)
                if any(
                    column <= len(rows[index - 1]) and bool(_string(rows[index - 1][column - 1]))
                    for column in range(column_start, column_end + 1)
                )
            ]
            if populated_rows:
                candidates.append(
                    Region(populated_rows[0], populated_rows[-1], column_start, column_end)
                )
    return candidates or [Region(1, max(len(rows), 1), 1, 1)]


def _slice_region(rows: list[list[Any]], region: Region) -> list[list[Any]]:
    return [
        row[region.start_column - 1 : region.end_column]
        for row in rows[region.start_row - 1 : region.end_row]
    ]


def _header_score(row: list[Any], following: list[Any] | None) -> tuple[float, list[str]]:
    values = [_string(value) for value in row]
    nonempty = [value for value in values if value]
    if len(nonempty) < 2:
        return 0.0, ["fewer than two non-empty cells"]
    unique_ratio = len(set(value.casefold() for value in nonempty)) / len(nonempty)
    text_ratio = sum(not INTEGER.fullmatch(value) and not DECIMAL.fullmatch(value) for value in nonempty) / len(
        nonempty
    )
    density = len(nonempty) / max(len(values), 1)
    next_density = (
        0.0
        if following is None
        else sum(bool(_string(value)) for value in following) / max(len(following), 1)
    )
    score = min(1.0, 0.35 * density + 0.3 * unique_ratio + 0.25 * text_ratio + 0.1 * next_density)
    return score, [
        f"{len(nonempty)} populated labels",
        f"{unique_ratio:.0%} unique labels",
        f"{text_ratio:.0%} text-like labels",
    ]


def _dedupe_headers(labels: list[str]) -> list[str]:
    headers: list[str] = []
    seen: Counter[str] = Counter()
    for index, value in enumerate(labels, start=1):
        base = value or f"unnamed_{index}"
        seen[base.casefold()] += 1
        headers.append(base if seen[base.casefold()] == 1 else f"{base}_{seen[base.casefold()]}")
    return headers


def _flatten_header_rows(rows: list[list[Any]], separator: str) -> list[str]:
    if not rows:
        return []
    width = max(len(row) for row in rows)
    levels: list[list[str]] = []
    for row in rows:
        forwarded: list[str] = []
        previous = ""
        for column in range(width):
            value = _string(row[column]) if column < len(row) else ""
            if value:
                previous = value
            forwarded.append(value or previous)
        levels.append(forwarded)
    flattened: list[str] = []
    for column in range(width):
        parts: list[str] = []
        for level in levels:
            value = level[column]
            if value and (not parts or value.casefold() != parts[-1].casefold()):
                parts.append(value)
        flattened.append(separator.join(parts))
    return _dedupe_headers(flattened)


def _candidate_headers(
    rows: list[list[Any]], absolute_start_row: int, depth: int, max_levels: int, separator: str
) -> list[HeaderCandidate]:
    candidates: list[HeaderCandidate] = []
    search_rows = rows[:depth]
    for index, row in enumerate(search_rows):
        following = search_rows[index + 1] if index + 1 < len(search_rows) else None
        single_score, evidence = _header_score(row, following)
        if single_score >= 0.35:
            labels = [_string(value) for value in row]
            flattened = _flatten_header_rows([row], separator)
            candidates.append(
                HeaderCandidate(
                    row_number=absolute_start_row + index,
                    row_numbers=[absolute_start_row + index],
                    confidence=round(single_score, 3),
                    labels=labels,
                    flattened_labels=flattened,
                    evidence=evidence + ["single-level header candidate"],
                )
            )
        for level_count in range(2, max_levels + 1):
            group = search_rows[index : index + level_count]
            if len(group) != level_count:
                continue
            first_values = [_string(value) for value in group[0]]
            first_present = [value for value in first_values if value]
            structural = len(first_present) >= 2 and (
                len(first_present) < len(first_values)
                or len({value.casefold() for value in first_present}) < len(first_present)
            )
            if not structural:
                continue
            scores = [
                _header_score(item, group[offset + 1] if offset + 1 < len(group) else None)[0]
                for offset, item in enumerate(group)
            ]
            flattened = _flatten_header_rows(group, separator)
            present_flattened = [value for value in flattened if value]
            unique_ratio = (
                len({value.casefold() for value in present_flattened}) / len(present_flattened)
                if present_flattened
                else 0
            )
            confidence = min(1.0, sum(scores) / len(scores) + 0.08 * (level_count - 1) + 0.08 * unique_ratio)
            if confidence >= 0.5:
                row_numbers = [absolute_start_row + index + offset for offset in range(level_count)]
                candidates.append(
                    HeaderCandidate(
                        row_number=row_numbers[0],
                        row_numbers=row_numbers,
                        confidence=round(confidence, 3),
                        labels=[_string(value) for value in group[-1]],
                        flattened_labels=flattened,
                        evidence=[
                            f"{level_count}-level header structure",
                            "blank or repeated parent labels support grouped columns",
                            f"{unique_ratio:.0%} unique flattened labels",
                        ],
                    )
                )
    return sorted(candidates, key=lambda item: (-item.confidence, item.row_number, len(item.row_numbers)))[:12]


def _infer_type(values: list[str]) -> tuple[CanonicalType, list[str]]:
    present = [value for value in values if value.casefold() not in NULL_LIKE]
    warnings: list[str] = []
    if not present:
        return CanonicalType.TEXT, warnings
    if any(DATE_AMBIGUOUS.fullmatch(value) for value in present):
        warnings.append("Ambiguous date-like values were left as text")
    if all(INTEGER.fullmatch(value) for value in present):
        if any(len(value.lstrip("+-")) > 1 and value.lstrip("+-").startswith("0") for value in present):
            warnings.append("Leading-zero values retained as identifier text")
            return CanonicalType.TEXT, warnings
        return CanonicalType.INTEGER, warnings
    if all(INTEGER.fullmatch(value) or DECIMAL.fullmatch(value) for value in present):
        return CanonicalType.DECIMAL, warnings
    if all(DATE_ISO.fullmatch(value) for value in present):
        return CanonicalType.DATE, warnings
    lowered = {value.casefold() for value in present}
    if lowered <= {"true", "false", "yes", "no"}:
        return CanonicalType.BOOLEAN, warnings
    numeric_count = sum(bool(INTEGER.fullmatch(value) or DECIMAL.fullmatch(value)) for value in present)
    if 0 < numeric_count < len(present):
        warnings.append("Mixed text and numeric values were left as text")
    return CanonicalType.TEXT, warnings


def _semantic_roles(name: str, values: list[str], inferred: CanonicalType) -> list[str]:
    roles: list[str] = []
    token = name.casefold()
    if any(word in token for word in ("id", "code", "number", "no", "ref")):
        roles.append("identifier")
    if inferred in {CanonicalType.DATE, CanonicalType.DATETIME} or "date" in token:
        roles.append("date")
    if inferred == CanonicalType.DECIMAL or any(word in token for word in ("amount", "value", "price", "total")):
        roles.append("amount")
    if "percent" in token or "pct" in token or any(PERCENT.fullmatch(value) for value in values):
        roles.append("percentage")
    if any(word in token for word in ("status", "category", "type")):
        roles.append("category")
    return list(dict.fromkeys(roles))


def _profile(headers: list[str], rows: list[list[Any]], sample_limit: int) -> list[ColumnProfile]:
    sampled = rows[:sample_limit]
    result: list[ColumnProfile] = []
    row_count = len(sampled)
    for index, name in enumerate(headers):
        values = [_string(row[index]) if index < len(row) else "" for row in sampled]
        present = [value for value in values if value.casefold() not in NULL_LIKE]
        counts = Counter(present)
        inferred, warnings = _infer_type(values)
        if len(rows) > len(sampled):
            warnings.append(f"Profile sampled {len(sampled)} of {len(rows)} rows")
        roles = _semantic_roles(name, present[:100], inferred)
        unique_count = len(counts)
        duplicate_count = sum(count - 1 for count in counts.values() if count > 1)
        identifier = "identifier" in roles or any(len(value) > 1 and value.startswith("0") for value in present[:100])
        result.append(
            ColumnProfile(
                source_name=name,
                inferred_type=inferred,
                null_percentage=round((row_count - len(present)) * 100 / row_count, 2) if row_count else 0,
                unique_count=unique_count,
                duplicate_count=duplicate_count,
                sample_values=list(dict.fromkeys(present))[:5],
                semantic_roles=roles,
                is_identifier_candidate=identifier,
                is_key_candidate=bool(present) and unique_count == len(present) and len(present) == row_count,
                warnings=warnings,
            )
        )
    return result


def _normalised(values: list[Any]) -> list[str]:
    return [_string(value).casefold() for value in values]


def _classify_rows(
    rows: list[list[Any]], absolute_start: int, flattened_headers: list[str], leaf_headers: list[str]
) -> list[RowClassification]:
    classifications: list[RowClassification] = []
    flattened_key = [value.casefold() for value in flattened_headers]
    leaf_key = [value.casefold() for value in _dedupe_headers(leaf_headers)]
    for offset, row in enumerate(rows):
        row_number = absolute_start + offset
        normalised = _normalised(row[: len(flattened_headers)])
        populated = [value for value in normalised if value]
        joined = " ".join(populated)
        kind: RowKind = "data"
        confidence = 1.0
        evidence = ["row retained as table data"]
        if normalised in (flattened_key, leaf_key):
            kind, confidence, evidence = "repeated_header", 1.0, ["row labels match selected header"]
        elif re.search(r"\bgrand\s+total\b", joined):
            kind, confidence, evidence = "grand_total", 0.98, ["grand total label detected"]
        elif re.search(r"\bsub\s*total\b", joined):
            kind, confidence, evidence = "subtotal", 0.95, ["subtotal label detected"]
        elif re.search(r"\b(generated|prepared|printed)\s+by\b", joined):
            kind, confidence, evidence = "generated_footer", 0.94, ["generated/prepared footer phrase detected"]
        elif re.search(r"\b(signature|approved by|authorised by)\b", joined):
            kind, confidence, evidence = "signature", 0.9, ["signature/approval phrase detected"]
        elif populated and (joined.startswith("note:") or joined.startswith("notes:")):
            kind, confidence, evidence = "note", 0.9, ["note prefix detected"]
        classifications.append(
            RowClassification(
                row_number=row_number,
                classification=kind,
                confidence=confidence,
                evidence=evidence,
            )
        )
    return classifications


def _discover_region(sheet: RawSheet, region: Region, overrides: DiscoveryOverrides) -> TableDiscovery:
    region_rows = _slice_region(sheet.rows, region)
    candidates = _candidate_headers(
        region_rows,
        region.start_row,
        overrides.header_search_depth,
        overrides.max_header_levels,
        overrides.header_flattening_separator,
    )
    override_rows = overrides.header_rows or ([overrides.header_row] if overrides.header_row else None)
    if override_rows:
        selected_rows = override_rows
        local_rows = [
            region_rows[row_number - region.start_row]
            for row_number in selected_rows
            if region.start_row <= row_number <= region.end_row
        ]
        if len(local_rows) != len(selected_rows):
            raise ValueError(f"HEADER_OVERRIDE_OUTSIDE_TABLE: {selected_rows}")
        headers = _flatten_header_rows(local_rows, overrides.header_flattening_separator)
        selected_confidence = 1.0
    elif candidates:
        selected_candidate = candidates[0]
        selected_rows = selected_candidate.row_numbers or [selected_candidate.row_number]
        headers = selected_candidate.flattened_labels or _dedupe_headers(selected_candidate.labels)
        selected_confidence = selected_candidate.confidence
    else:
        selected_rows = [region.start_row]
        headers = _dedupe_headers([_string(value) for value in region_rows[0]]) if region_rows else []
        selected_confidence = 0.0
    data_start = selected_rows[-1] + 1
    local_data_start = max(0, data_start - region.start_row)
    raw_data_rows = region_rows[local_data_start:]
    leaf_labels = (
        [_string(value) for value in region_rows[selected_rows[-1] - region.start_row]]
        if region_rows and selected_rows[-1] <= region.end_row
        else []
    )
    classifications = _classify_rows(raw_data_rows, data_start, headers, leaf_labels)
    data_rows = [
        row
        for row, classification in zip(raw_data_rows, classifications, strict=True)
        if classification.classification == "data"
    ]
    repeated = [item.row_number for item in classifications if item.classification == "repeated_header"]
    footer_rows = [
        item.row_number
        for item in classifications
        if item.classification not in {"data", "repeated_header"}
    ]
    profiles = _profile(headers, data_rows, overrides.profile_sample_rows)
    preview = [
        dict(zip(headers, [_string(value) for value in row], strict=False))
        for row in data_rows[: overrides.preview_rows]
    ]
    warnings: list[str] = []
    if not candidates:
        warnings.append("No confident header candidate was found; user confirmation is required")
    if repeated:
        warnings.append(f"Detected {len(repeated)} repeated header row(s); no row was deleted")
    if footer_rows:
        warnings.append(f"Classified {len(footer_rows)} footer/total/note row(s); no row was deleted")
    if len(candidates) > 1 and candidates[0].confidence - candidates[1].confidence < 0.08:
        warnings.append("Header selection is ambiguous")
    cell_count = (region.end_row - region.start_row + 1) * (region.end_column - region.start_column + 1)
    if cell_count > 5_000_000:
        warnings.append("Table exceeds the default 5,000,000-cell memory-risk warning threshold")
    merged_in_header = [
        item
        for item in sheet.merged_ranges
        if item[0] <= selected_rows[-1]
        and item[1] >= selected_rows[0]
        and item[2] <= region.end_column
        and item[3] >= region.start_column
    ]
    evidence = [
        f"non-empty rectangular region {region.region_name}",
        f"selected header rows {selected_rows}",
        f"{len(data_rows)} probable data rows",
    ]
    if merged_in_header:
        evidence.append(f"{len(merged_in_header)} merged header range(s) forward-filled")
    table_id = (
        f"sheet:{sheet.name}:r{region.start_row}-{region.end_row}:"
        f"c{region.start_column}-{region.end_column}"
    )
    return TableDiscovery(
        table_id=table_id,
        sheet_name=sheet.name,
        sheet_state=sheet.state,
        candidate_region=region.region_name,
        candidate_headers=candidates,
        selected_header_row=selected_rows[0],
        selected_header_rows=selected_rows,
        header_flattening_separator=overrides.header_flattening_separator,
        start_row=region.start_row,
        end_row=region.end_row,
        start_column=region.start_column,
        end_column=region.end_column,
        row_count_estimate=len(data_rows),
        column_count=len(headers),
        blank_leading_rows=region.start_row - 1,
        blank_trailing_rows=max(0, len(sheet.rows) - region.end_row),
        repeated_header_rows=repeated,
        footer_rows=footer_rows,
        row_classifications=classifications,
        columns=profiles,
        preview=preview,
        confidence=selected_confidence,
        decision=f"selected {region.region_name} with header rows {selected_rows}",
        evidence=evidence,
        user_override={
            "table_id": overrides.table_id,
            "sheet_name": overrides.sheet_name,
            "header_rows": override_rows,
            "header_flattening_separator": overrides.header_flattening_separator,
        },
        warnings=warnings,
    )


def _sheet_discoveries(sheet: RawSheet, overrides: DiscoveryOverrides) -> list[TableDiscovery]:
    tables = [_discover_region(sheet, region, overrides) for region in _regions(sheet.rows)]
    table_ids = [table.table_id for table in tables]
    for table in tables:
        table.alternative_candidates = [table_id for table_id in table_ids if table_id != table.table_id]
    if overrides.table_id:
        tables = [table for table in tables if table.table_id == overrides.table_id]
        if not tables:
            raise ValueError(f"TABLE_NOT_FOUND: {overrides.table_id}")
    return tables


def discover_source(source: SourceFile, handle: SourceHandle, overrides: DiscoveryOverrides) -> DiscoveryResult:
    source.assert_unchanged()
    sheets = load_sheets(source)
    selected_sheets = [sheet for sheet in sheets if overrides.sheet_name in (None, sheet.name)]
    if overrides.sheet_name and not selected_sheets:
        raise ValueError(f"SHEET_NOT_FOUND: {overrides.sheet_name}")
    tables = [table for sheet in selected_sheets for table in _sheet_discoveries(sheet, overrides)]
    warnings = [] if tables else ["No sheets or tables were found"]
    source.assert_unchanged()
    return DiscoveryResult(source=handle, tables=tables, warnings=warnings)


def read_selected_table(source: SourceFile, overrides: DiscoveryOverrides, limit: int | None = None) -> pl.DataFrame:
    sheets = (
        _load_csv(
            source.path,
            overrides.header_search_depth + overrides.max_header_levels + limit,
        )
        if source.path.suffix.lower() == ".csv" and limit is not None
        else load_sheets(source)
    )
    sheet = next((item for item in sheets if overrides.sheet_name in (None, item.name)), None)
    if sheet is None:
        raise ValueError("SHEET_NOT_FOUND")
    discoveries = _sheet_discoveries(sheet, overrides)
    discovery = discoveries[0] if discoveries else None
    if discovery is None or not discovery.columns:
        return pl.DataFrame()
    headers = [column.source_name for column in discovery.columns]
    rows = [
        row[discovery.start_column - 1 : discovery.end_column]
        for row in sheet.rows[discovery.selected_header_rows[-1] : discovery.end_row]
    ]
    if limit is not None:
        rows = rows[:limit]
    records = [
        {header: _string(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
        for row in rows
    ]
    return pl.DataFrame(records, schema={header: pl.String for header in headers})
