"""Semantic workbook signatures that ignore volatile XLSX ZIP metadata."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def workbook_signature(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            headers = [cell.value for cell in sheet[1]]
            types = sorted(
                {
                    cell.data_type
                    for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 100))
                    for cell in row
                    if cell.value is not None
                }
            )
            number_formats = sorted(
                {
                    cell.number_format
                    for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 100))
                    for cell in row
                    if cell.value is not None
                }
            )
            formula_cells = [
                cell.coordinate
                for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 100))
                for cell in row
                if cell.data_type == "f"
            ]
            sheets.append(
                {
                    "name": sheet.title,
                    "headers": headers,
                    "max_column": sheet.max_column,
                    "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                    "auto_filter": sheet.auto_filter.ref,
                    "data_types": types,
                    "number_formats": number_formats,
                    "formula_cells": formula_cells,
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def compare_workbook_structure(expected: Path, actual: Path) -> list[str]:
    left = workbook_signature(expected)
    right = workbook_signature(actual)
    differences: list[str] = []
    left_sheets = {item["name"]: item for item in left["sheets"]}
    right_sheets = {item["name"]: item for item in right["sheets"]}
    if list(left_sheets) != list(right_sheets):
        differences.append(
            f"sheet order/names differ: expected {list(left_sheets)}, observed {list(right_sheets)}"
        )
    for name in left_sheets.keys() & right_sheets.keys():
        for field in (
            "headers",
            "max_column",
            "freeze_panes",
            "auto_filter",
            "data_types",
            "number_formats",
            "formula_cells",
        ):
            if left_sheets[name][field] != right_sheets[name][field]:
                differences.append(
                    f"{name}.{field} differs: expected {left_sheets[name][field]!r}, "
                    f"observed {right_sheets[name][field]!r}"
                )
    return differences
