from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from packages.contracts import OperationNode, SourceHandle, WorkflowConfiguration
from packages.data_engine import EngineRuntime, RuntimeExecutionError, Workspace
from packages.data_engine.exporter import SHEETS
from packages.data_engine.safety import sha256_file


def test_vertical_slice_exports_audited_workbook_without_changing_source(
    fixture_dir: Path, tmp_path: Path, workflow: WorkflowConfiguration
) -> None:
    workspace = Workspace(tmp_path / "runtime")
    original = fixture_dir / "header_row_1.csv"
    original_before = sha256_file(original)
    source = workspace.import_source(original, original.name)
    handle = SourceHandle(
        id=source.id,
        project_id=workflow.project_id,
        original_filename=original.name,
        media_type="text/csv",
        size_bytes=source.size_bytes,
        sha256=source.sha256,
    )
    result = EngineRuntime(workspace).execute(source, handle, workflow)
    assert result.record.rows_read == 4
    assert result.record.rows_written + result.record.rows_rejected + result.record.rows_filtered == 4
    assert result.record.rows_rejected == 3
    assert sha256_file(original) == original_before
    source.assert_unchanged()
    workbook_path = Path(result.record.artifacts[0])
    workbook = load_workbook(workbook_path, read_only=True)
    try:
        assert set(SHEETS) <= set(workbook.sheetnames)
        summary = workbook["Summary"]
        values = [cell.value for row in summary.iter_rows() for cell in row]
        assert source.sha256 in values
        assert str(result.record.id) in values
    finally:
        workbook.close()


def test_saved_workflow_runs_with_reordered_columns(
    fixture_dir: Path, tmp_path: Path, workflow: WorkflowConfiguration
) -> None:
    workspace = Workspace(tmp_path / "runtime")
    source = workspace.import_source(fixture_dir / "reordered_columns.csv", "reordered_columns.csv")
    preview = EngineRuntime(workspace).preview(source, workflow)
    assert preview.rows_read == 4
    assert preview.rows[0]["employee_id"] == "00124"


def test_failed_execution_produces_failed_audit_evidence(
    fixture_dir: Path, tmp_path: Path, workflow: WorkflowConfiguration
) -> None:
    workspace = Workspace(tmp_path / "runtime")
    source = workspace.import_source(fixture_dir / "header_row_1.csv", "header_row_1.csv")
    handle = SourceHandle(
        id=source.id,
        project_id=workflow.project_id,
        original_filename=source.original_filename,
        media_type="text/csv",
        size_bytes=source.size_bytes,
        sha256=source.sha256,
    )
    broken = workflow.model_copy(
        update={"operations": [OperationNode(operation_id="processor.not_installed")]}
    )
    with pytest.raises(RuntimeExecutionError) as captured:
        EngineRuntime(workspace).execute(source, handle, broken)
    assert captured.value.record.status == "failed"
    assert captured.value.record.artifacts
    assert Path(captured.value.record.artifacts[0]).is_file()
    source.assert_unchanged()
