from __future__ import annotations

import threading
import time
from datetime import UTC, datetime
from pathlib import Path
from uuid import UUID, uuid4

import polars as pl
import pytest
from openpyxl import load_workbook

from apps.api.app.dag_repository import SQLiteDagRepository
from apps.api.app.database import Database
from packages.contracts import (
    ArtifactType,
    DagEdge,
    DagJobSubmission,
    DagNode,
    DagOutputDefinition,
    DagPosition,
    DagRunRecord,
    DagRunRequest,
    DagRunStatus,
    DagWorkflow,
    ManualCheckpointDecision,
    SubflowDefinition,
    WorkflowLifecycle,
)
from packages.workflow_dag.adapters import DagAdapterRegistry, NodeInputs, RuntimeControl, engine_adapter_registry
from packages.workflow_dag.planner import build_execution_plan
from packages.workflow_dag.registry import default_registry
from packages.workflow_dag.runtime import LocalDagExecutor


def _node(node_id: str, type_id: str, x: float = 0) -> DagNode:
    capability = default_registry.require(type_id, 1)
    configuration = {"source_id": str(uuid4())} if type_id == "source.saved_dataset" else {}
    return DagNode(
        id=node_id,
        node_type_id=type_id,
        display_name=node_id.title(),
        category=capability.category,
        position=DagPosition(x=x, y=0),
        configuration=configuration,
        input_ports=capability.input_ports,
        output_ports=capability.output_ports,
        retry_classification=capability.retry_classification,
        entitlement_capability_id=capability.entitlement_capability_id,
    )


def _store(tmp_path: Path) -> tuple[SQLiteDagRepository, UUID]:
    database = Database(tmp_path / "metadata.sqlite3")
    database.initialize()
    project_id = uuid4()
    now = datetime.now(UTC).isoformat()
    with database.connect() as connection:
        connection.execute(
            "INSERT INTO projects VALUES (?, ?, ?, ?, ?, ?)",
            (str(project_id), "DAG test", "en-IN", "local_only", now, now),
        )
    return SQLiteDagRepository(database), project_id


def _wait(store: SQLiteDagRepository, run_id: UUID, statuses: set[DagRunStatus]) -> DagRunRecord:
    for _ in range(100):
        record = store.get_run(run_id)
        if record is not None and record.status in statuses:
            return record
        time.sleep(0.02)
    raise AssertionError("DAG run did not reach expected state")


def test_background_dag_materialises_artifacts_and_succeeds(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    source = _node("source", "source.saved_dataset")
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Background source",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[source],
        outputs=[
            DagOutputDefinition(
                id="dataset",
                display_name="Dataset",
                node_id="source",
                port_id="dataset",
                artifact_type=ArtifactType.CANONICAL_DATASET,
            )
        ],
    )
    adapters = DagAdapterRegistry()

    def load(_: DagNode, __: NodeInputs, control: RuntimeControl) -> dict[str, object]:
        control.check_cancelled()
        return {"dataset": pl.DataFrame({"record_id": ["A", "B"], "amount": [10, 20]})}

    adapters.register("source.saved_dataset", load)
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        completed = _wait(store, queued.id, {DagRunStatus.SUCCEEDED})
        assert completed.output_available
        assert completed.progress_percent == 100
        node_runs = store.list_node_runs(queued.id)
        assert node_runs[0].status.value == "succeeded"
        artifact = node_runs[0].output_artifacts[0]
        assert artifact.row_count == 2
        assert artifact.sha256
        assert (tmp_path / "workspace" / "dag-runs" / "completed" / str(queued.id) / artifact.path_reference).is_file()
        with pytest.raises(ValueError, match="DAG_RUN_NOT_RESUMABLE"):
            executor.resume(queued.id)
    finally:
        executor.shutdown()


def test_manual_checkpoint_waits_for_immutable_decision_then_resumes(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    source = _node("source", "source.saved_dataset")
    review = _node("review", "control.manual_approval", 300)
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Governed publication",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[source, review],
        edges=[
            DagEdge(
                id="source_review",
                source_node_id="source",
                source_port_id="dataset",
                target_node_id="review",
                target_port_id="input",
                data_contract_reference="canonical-dataset/v1",
            )
        ],
        outputs=[
            DagOutputDefinition(
                id="approved",
                display_name="Approved dataset",
                node_id="review",
                port_id="approved",
                artifact_type=ArtifactType.ANY,
            )
        ],
    )
    adapters = DagAdapterRegistry()
    adapters.register(
        "source.saved_dataset",
        lambda node, inputs, control: {"dataset": pl.DataFrame({"key": ["A"]})},
    )
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        waiting = _wait(store, queued.id, {DagRunStatus.WAITING_FOR_REVIEW})
        assert not waiting.output_available
        checkpoint = store.list_checkpoints(queued.id)[0]
        decision = ManualCheckpointDecision(
            checkpoint_id=checkpoint.id,
            action="approve",
            actor="test-reviewer",
            comment="Evidence reviewed",
        )
        store.append_decision(decision)
        executor.resume(queued.id)
        completed = _wait(store, queued.id, {DagRunStatus.SUCCEEDED})
        assert completed.output_available
        decisions = store.list_decisions(checkpoint.id)
        assert decisions == [decision]
    finally:
        executor.shutdown()


def test_orphaned_running_dag_requires_explicit_recovery(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    source = _node("source", "source.saved_dataset")
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Interrupted workflow",
        nodes=[source],
        outputs=[
            DagOutputDefinition(
                id="dataset",
                display_name="Dataset",
                node_id="source",
                port_id="dataset",
                artifact_type=ArtifactType.CANONICAL_DATASET,
            )
        ],
    )
    store.save_workflow(workflow)
    actual_plan = build_execution_plan(workflow)
    store.save_plan(actual_plan)
    run = DagRunRecord(
        project_id=workflow.project_id,
        workflow_id=workflow.id,
        workflow_version=workflow.version,
        plan_id=actual_plan.id,
        status=DagRunStatus.RUNNING,
    )
    store.create_run(run, DagRunRequest(workflow=workflow))
    recovered = store.recover_orphans()
    assert recovered[0].status.value == "recovery_required"
    assert not recovered[0].output_available


def test_independent_ready_sources_execute_in_parallel(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    left = _node("left", "source.saved_dataset")
    right = _node("right", "source.saved_dataset")
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Parallel ready nodes",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[left, right],
        outputs=[
            DagOutputDefinition(
                id="left_output",
                display_name="Left",
                node_id="left",
                port_id="dataset",
                artifact_type=ArtifactType.CANONICAL_DATASET,
            ),
            DagOutputDefinition(
                id="right_output",
                display_name="Right",
                node_id="right",
                port_id="dataset",
                artifact_type=ArtifactType.CANONICAL_DATASET,
            ),
        ],
    )
    barrier = threading.Barrier(2)
    adapters = DagAdapterRegistry()

    def concurrent_source(node: DagNode, inputs: NodeInputs, control: RuntimeControl) -> dict[str, object]:
        del inputs
        control.check_cancelled()
        barrier.wait(timeout=2)
        return {"dataset": pl.DataFrame({"node": [node.id]})}

    adapters.register("source.saved_dataset", concurrent_source)
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        completed = _wait(store, queued.id, {DagRunStatus.SUCCEEDED, DagRunStatus.FAILED})
        assert completed.status == DagRunStatus.SUCCEEDED
        assert sorted(completed.completed_node_ids) == ["left", "right"]
    finally:
        executor.shutdown()


def test_conditional_route_records_inactive_branch_as_skipped(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    source = _node("source", "source.saved_dataset")
    condition = _node("condition", "control.condition", 250)
    condition.configuration = {
        "kind": "literal",
        "value": True,
        "value_type": "boolean",
        "field_id": None,
        "function": None,
        "args": [],
    }
    true_path = _node("true_path", "control.merge", 500)
    true_path.configuration = {"strategy": "first_available"}
    false_path = _node("false_path", "control.merge", 500)
    false_path.configuration = {"strategy": "first_available"}
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Conditional route",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[source, condition, true_path, false_path],
        edges=[
            DagEdge(
                id="source_condition",
                source_node_id="source",
                source_port_id="dataset",
                target_node_id="condition",
                target_port_id="input",
                data_contract_reference="any/v1",
            ),
            DagEdge(
                id="condition_true",
                source_node_id="condition",
                source_port_id="true",
                target_node_id="true_path",
                target_port_id="branches",
                data_contract_reference="control/v1",
            ),
            DagEdge(
                id="condition_false",
                source_node_id="condition",
                source_port_id="false",
                target_node_id="false_path",
                target_port_id="branches",
                data_contract_reference="control/v1",
            ),
        ],
        outputs=[
            DagOutputDefinition(
                id="true_output",
                display_name="True",
                node_id="true_path",
                port_id="output",
                artifact_type=ArtifactType.ANY,
            ),
            DagOutputDefinition(
                id="false_output",
                display_name="False",
                node_id="false_path",
                port_id="output",
                artifact_type=ArtifactType.ANY,
            ),
        ],
    )
    adapters = engine_adapter_registry()
    adapters.register("source.saved_dataset", lambda node, inputs, control: {"dataset": pl.DataFrame({"value": [1]})})
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        completed = _wait(store, queued.id, {DagRunStatus.SUCCEEDED, DagRunStatus.FAILED})
        assert completed.status == DagRunStatus.SUCCEEDED
        assert completed.skipped_node_ids == ["false_path"]
        by_node = {record.node_id: record for record in store.list_node_runs(queued.id)}
        assert by_node["true_path"].status.value == "succeeded"
        assert by_node["false_path"].status.value == "skipped"
    finally:
        executor.shutdown()


def test_dag_excel_output_is_formula_injection_safe(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    source = _node("source", "source.saved_dataset")
    output = _node("excel", "output.excel", 300)
    output.configuration = {
        "filename_prefix": "formula_safe",
        "include_summary": True,
        "include_rejected_rows": True,
        "include_source_metadata": True,
    }
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Formula-safe output",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[source, output],
        edges=[
            DagEdge(
                id="source_excel",
                source_node_id="source",
                source_port_id="dataset",
                target_node_id="excel",
                target_port_id="input",
                data_contract_reference="canonical_dataset/v1",
            )
        ],
        outputs=[
            DagOutputDefinition(
                id="package",
                display_name="Evidence",
                node_id="excel",
                port_id="package",
                artifact_type=ArtifactType.EVIDENCE_PACKAGE,
            )
        ],
    )
    adapters = engine_adapter_registry()
    adapters.register(
        "source.saved_dataset",
        lambda node, inputs, control: {"dataset": pl.DataFrame({"evidence": ["=2+2"]})},
    )
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        completed = _wait(store, queued.id, {DagRunStatus.SUCCEEDED, DagRunStatus.FAILED})
        assert completed.status == DagRunStatus.SUCCEEDED
        record = next(item for item in store.list_node_runs(queued.id) if item.node_id == "excel")
        artifact = record.output_artifacts[0]
        path = tmp_path / "workspace" / "dag-runs" / "completed" / str(queued.id) / artifact.path_reference
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            assert worksheet["A2"].value == "'=2+2"
            assert worksheet["A2"].data_type != "f"
        finally:
            workbook.close()
    finally:
        executor.shutdown()


def test_cancellation_and_deterministic_retry_have_terminal_truthful_states(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    source = _node("source", "source.saved_dataset")
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Cancellation and retry",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[source],
        outputs=[
            DagOutputDefinition(
                id="dataset",
                display_name="Dataset",
                node_id="source",
                port_id="dataset",
                artifact_type=ArtifactType.CANONICAL_DATASET,
            )
        ],
    )
    started = threading.Event()
    adapters = DagAdapterRegistry()

    def cancellable(node: DagNode, inputs: NodeInputs, control: RuntimeControl) -> dict[str, object]:
        del node, inputs
        started.set()
        for _ in range(100):
            time.sleep(0.01)
            control.check_cancelled()
        return {"dataset": pl.DataFrame({"value": [1]})}

    adapters.register("source.saved_dataset", cancellable)
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        assert started.wait(timeout=2)
        executor.cancel(queued.id)
        cancelled = _wait(store, queued.id, {DagRunStatus.CANCELLED, DagRunStatus.FAILED})
        assert cancelled.status == DagRunStatus.CANCELLED
        assert not cancelled.output_available
    finally:
        executor.shutdown()

    retry_adapters = DagAdapterRegistry()
    should_fail = True

    def deterministic(node: DagNode, inputs: NodeInputs, control: RuntimeControl) -> dict[str, object]:
        del node, inputs
        control.check_cancelled()
        if should_fail:
            raise ValueError("SYNTHETIC_DETERMINISTIC_FAILURE")
        return {"dataset": pl.DataFrame({"value": [1]})}

    retry_adapters.register("source.saved_dataset", deterministic)
    retry_executor = LocalDagExecutor(store, retry_adapters, tmp_path / "retry-workspace")
    try:
        failed_run = retry_executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        failed = _wait(store, failed_run.id, {DagRunStatus.FAILED})
        assert not failed.output_available
        should_fail = False
        retried = retry_executor.retry(failed.id)
        completed = _wait(store, retried.id, {DagRunStatus.SUCCEEDED, DagRunStatus.FAILED})
        assert completed.status == DagRunStatus.SUCCEEDED
        assert completed.retry_of == failed.id
    finally:
        retry_executor.shutdown()


def test_version_pinned_subflow_executes_with_namespaced_lineage(tmp_path: Path) -> None:
    store, project_id = _store(tmp_path)
    inner_source = _node("inner_source", "source.saved_dataset")
    subflow = SubflowDefinition(
        project_id=project_id,
        display_name="Reusable source subflow",
        public_input_ports=[],
        public_output_ports=inner_source.output_ports,
        nodes=[inner_source],
        edges=[],
    )
    store.save_subflow(subflow)
    reference = _node("subflow", "subflow.reference")
    reference.configuration = {
        "subflow_id": str(subflow.id),
        "subflow_version": 1,
        "input_bindings": {},
        "output_bindings": {"output": "inner_source.dataset"},
    }
    workflow = DagWorkflow(
        project_id=project_id,
        display_name="Pinned subflow execution",
        lifecycle=WorkflowLifecycle.PUBLISHED,
        nodes=[reference],
        outputs=[
            DagOutputDefinition(
                id="result",
                display_name="Result",
                node_id="subflow",
                port_id="output",
                artifact_type=ArtifactType.ANY,
            )
        ],
    )
    adapters = DagAdapterRegistry()
    adapters.register(
        "source.saved_dataset",
        lambda node, inputs, control: {"dataset": pl.DataFrame({"lineage": [node.id]})},
    )
    executor = LocalDagExecutor(store, adapters, tmp_path / "workspace")
    try:
        queued = executor.submit(DagJobSubmission(request=DagRunRequest(workflow=workflow)))
        completed = _wait(store, queued.id, {DagRunStatus.SUCCEEDED, DagRunStatus.FAILED})
        assert completed.status == DagRunStatus.SUCCEEDED
        assert completed.completed_node_ids == ["subflow.inner_source"]
        assert store.list_node_runs(queued.id)[0].node_id == "subflow.inner_source"
    finally:
        executor.shutdown()
