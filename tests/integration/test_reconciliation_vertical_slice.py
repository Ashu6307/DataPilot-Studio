from __future__ import annotations

import csv
import time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

import polars as pl
from openpyxl import load_workbook

from apps.api.app.database import Database
from apps.api.app.reconciliation_job_store import SQLiteReconciliationJobStore
from apps.api.app.repositories import SQLiteMetadataRepository
from apps.api.app.services import DataPilotService
from packages.contracts import (
    BlockingMethod,
    CandidateConstraint,
    ComparisonConfiguration,
    FieldComparisonRule,
    MatchMethod,
    MatchStage,
    NumericTolerance,
    NumericToleranceMode,
    ProjectCreate,
    ReconciliationJobSubmission,
    ReconciliationRunRequest,
    ReconciliationWorkflow,
    RunStatus,
)
from packages.data_engine.reconciliation import reconcile_datasets
from packages.data_engine.reconciliation_background import LocalReconciliationJobExecutor
from packages.data_engine.reconciliation_exporter import export_reconciliation_evidence
from packages.data_engine.safety import Workspace, sha256_file


def _upload(service: DataPilotService, project_id: UUID, name: str, payload: str):  # type: ignore[no-untyped-def]
    return service.import_source(project_id, name, "text/csv", BytesIO(payload.encode("utf-8")))


def test_reconciliation_runs_in_background_with_outputs_and_checkpoints(tmp_path: Path) -> None:
    database = Database(tmp_path / "metadata.sqlite3")
    database.initialize()
    repository = SQLiteMetadataRepository(database)
    service = DataPilotService(repository, Workspace(tmp_path / "workspace"))
    project = service.create_project(ProjectCreate(name="Reconciliation vertical slice"))
    left = _upload(
        service,
        project.id,
        "left.csv",
        "key,name,region,amount,note\nA,Alpha,north,100,=1+1\nB,Beta,south,200,left\n",
    )
    right = _upload(
        service,
        project.id,
        "right.csv",
        "key,name,region,amount,note\nA,Alpha,north,100,safe\nX,Beta,south,202,right\n",
    )
    comparison = ComparisonConfiguration(
        project_id=project.id,
        left_dataset_id=left.id,
        right_dataset_id=right.id,
        business_key_fields=["key"],
        compare_fields=["amount", "note"],
    )
    workflow = ReconciliationWorkflow(
        project_id=project.id,
        display_name="Generic exact and tolerance",
        left_dataset_id=left.id,
        right_dataset_id=right.id,
        comparison=comparison,
        comparison_fields=[FieldComparisonRule(field_id="note")],
        stages=[
            MatchStage(
                id="exact_key",
                name="Exact key",
                priority=1,
                left_key_fields=["key"],
                right_key_fields=["key"],
                method=MatchMethod.EXACT,
            ),
            MatchStage(
                id="amount_tolerance",
                name="Amount tolerance",
                priority=2,
                left_key_fields=["amount"],
                right_key_fields=["amount"],
                method=MatchMethod.NUMERIC_TOLERANCE,
                threshold=Decimal("0.5"),
                numeric_tolerances={
                    "amount": NumericTolerance(
                        mode=NumericToleranceMode.ABSOLUTE,
                        tolerance=Decimal("5"),
                    )
                },
                candidate_constraints=[
                    CandidateConstraint(
                        id="same_region",
                        method=BlockingMethod.EXACT,
                        left_field="region",
                        right_field="region",
                    )
                ],
            ),
        ],
    )
    service.save_reconciliation_workflow(workflow)
    store = SQLiteReconciliationJobStore(database)
    executor = LocalReconciliationJobExecutor(store, service.run_reconciliation_background)
    left_path = service.workspace.source_from_id(left.id, left.original_filename, left.sha256).path
    right_path = service.workspace.source_from_id(right.id, right.original_filename, right.sha256).path
    before = (sha256_file(left_path), sha256_file(right_path))
    try:
        job = executor.submit(ReconciliationJobSubmission(run=ReconciliationRunRequest(workflow=workflow)))
        deadline = time.monotonic() + 10
        current = store.get(job.id)
        while current is not None and current.status not in {
            RunStatus.SUCCEEDED,
            RunStatus.FAILED,
            RunStatus.CANCELLED,
        }:
            assert time.monotonic() < deadline
            time.sleep(0.01)
            current = store.get(job.id)
        assert current is not None
        assert current.status == RunStatus.SUCCEEDED, current.error_message
        assert current.output_available
        assert current.run_id is not None
        checkpoints = store.checkpoints(job.id)
        assert any(checkpoint.completed_stage == "exact_key" for checkpoint in checkpoints)
        assert checkpoints[-1].completed_stage == "reconciliation_completed"
        record = repository.get_reconciliation_run(current.run_id)
        assert record is not None
        assert record.summary.exact_matches == 1
        assert record.summary.tolerance_matches == 1
        manifest = repository.get_reconciliation_manifest(current.run_id)
        assert manifest is not None
        assert {entry.media_type for entry in manifest.entries} >= {
            "text/csv",
            "application/json",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        workbook_entry = next(entry for entry in manifest.entries if entry.relative_path.endswith(".xlsx"))
        stored_run = repository.get_run(current.run_id)
        assert stored_run is not None
        workbook_path = next(Path(path) for path in stored_run.artifacts if path.endswith(workbook_entry.relative_path))
        workbook = load_workbook(workbook_path, read_only=True)
        assert "Reconciliation Summary" in workbook.sheetnames
        workbook.close()
        assert before == (sha256_file(left_path), sha256_file(right_path))
    finally:
        executor.shutdown()


def test_evidence_zip_is_deterministic_and_formula_safe(tmp_path: Path) -> None:
    workflow = ReconciliationWorkflow(
        project_id=uuid4(),
        display_name="Evidence",
        left_dataset_id=uuid4(),
        right_dataset_id=uuid4(),
        comparison_fields=[FieldComparisonRule(field_id="note")],
        stages=[
            MatchStage(
                id="exact",
                name="Exact",
                priority=1,
                left_key_fields=["key"],
                right_key_fields=["key"],
                method=MatchMethod.EXACT,
            )
        ],
    )
    result = reconcile_datasets(
        pl.DataFrame({"key": ["A"], "note": ["=1+1"]}),
        pl.DataFrame({"key": ["A"], "note": ["safe"]}),
        workflow,
    )
    first, second = tmp_path / "first", tmp_path / "second"
    export_reconciliation_evidence(first, result, workflow)
    export_reconciliation_evidence(second, result, workflow)
    assert (first / "reconciliation_evidence.zip").read_bytes() == (second / "reconciliation_evidence.zip").read_bytes()
    difference_csv = first / "csv" / "Field Differences.csv"
    assert "'=1+1" in difference_csv.read_text(encoding="utf-8-sig")
    with difference_csv.open(encoding="utf-8-sig", newline="") as stream:
        difference = next(csv.DictReader(stream))
    assert difference["left_dataset_id"] == str(workflow.left_dataset_id)
    assert difference["right_dataset_id"] == str(workflow.right_dataset_id)
    assert difference["stage_id"] == "exact"
    assert difference["workflow_version"] == "1"
    assert difference["run_id"] == str(result.run_id)
